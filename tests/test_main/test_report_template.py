"""Шаблоны отчетов (Модератор → Главное, biruni ker/template_list) — CRUD testlari.

MCP bilan real DOM'da tasdiqlangan:

Ro'yxat (template_list, heading "Шаблоны отчетов"): "Создать" + "Доступные шаблоны"
sub-nav. Ustunlar: Название / Название формы / Типы источника / Файл / Создал / Дата /
Статус ("active"/"passive" — inglizcha, _STATUS_SYNONYMS qoplaydi).

Create/Edit forma (heading IKKALASIDA "Настройки шаблонов"; create URL ker/setting+add,
edit ker/setting+edit):
  - "Источник *"      : smt-data-select — MAJBURIY (prod'da yagona variant "Накладная (заказ)")
  - "Название *"      : smt-input — MAJBURIY
  - "Файл шаблона *"  : input[type=file] accept=.xlsx — MAJBURIY (set_input_files bilan yuklanadi;
    EDIT'da fayl SAQLANADI — qayta yuklash shart emas)
  - "Статус"          : smt-switch (Активный)
  - ixtiyoriy: Группа шаблона, Тип шаблона (EXCEL/Word), Порядковый номер, 3 checkbox.
Save create'da template_list'ga REDIRECT qiladi.

Qator paneli: Изменить / Удалить / Прикрепить роли / Прикрепить пользователей.
DIQQAT: Просмотр (view) va Дубликат (duplicate) YO'Q — shu sabab view/duplicate testi yozilmadi.
Delete → cdk-overlay "Да/Нет" ("Удалить {nom}?"). Negativ: bo'sh forma + Сохранить →
formada qoladi (majburiy "*"), save bloklanadi.
"""
import os
import re
import tempfile

import allure
import openpyxl
import pytest
from playwright.sync_api import Page, expect

from flows.flow_authorization import authorization
from flows.flow_navbar import flow_navigate
from utils.base_page import BasePage

MENU = "Шаблоны отчетов"
SOURCE = "Накладная (заказ)"   # prod'da yagona mavjud Источник
FORM_HEADING = "Настройки шаблонов"   # create VA edit formasi sarlavhasi (bir xil)

_XLSX_PATH = None


def _template_xlsx() -> str:
    """Yuklash uchun vaqtinchalik .xlsx yaratadi (bir marta) va yo'lini qaytaradi.
    Fayl shabloni faqat .xlsx bo'lishi kerak — mazmuni muhim emas (server extension
    va tuzilishni qabul qiladi, MCP tasdiqlangan 2026-08-07)."""
    global _XLSX_PATH
    if _XLSX_PATH and os.path.exists(_XLSX_PATH):
        return _XLSX_PATH
    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="report_tpl_")
    os.close(fd)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Test"
    ws["A2"] = "#BREAK#"
    wb.save(path)
    _XLSX_PATH = path
    return path


def _open_create(page: Page, m: BasePage) -> None:
    flow_navigate(page, tab="Модератор", name=MENU, expect_url="template_list")
    m.expect_heading(MENU)
    m.open_create()
    m.expect_heading(FORM_HEADING)


def run_report_template(page: Page, code, name=None) -> dict:
    """Yangi Шаблон отчета yaratadi (Источник + Название + .xlsx fayl). ``name``
    berilmasa shablon-{code}. Yaratilgan qiymatlarni qaytaradi."""
    m = BasePage(page)
    if name is None:
        name = f"shablon-{code}"

    with allure.step(f"Навигация → Создать: {MENU}"):
        _open_create(page, m)

    with allure.step(f"Форма: Источник={SOURCE}, Название={name}, Файл шаблона=.xlsx"):
        m.select(SOURCE, label="Источник")
        m.input(label="Название", value=name)
        # Файл шаблона: yashirin input[type=file] (accept .xlsx) — to'g'ridan-to'g'ri
        # set_input_files (fayl tanlash dialogini ochmasdan). MCP tasdiqlangan 2026-08-07.
        page.locator('input[type=file]').set_input_files(_template_xlsx())

    with allure.step("Сохранить va ro'yxatda tekshirish"):
        m.save()  # create → template_list'ga redirect
        flow_navigate(page, tab="Модератор", name=MENU, expect_url="template_list")
        m.expect_heading(MENU)
        m.search(name)
        m.grid_row(name)

    return {"name": name}


@allure.epic("Модератор")
@allure.feature("Шаблоны отчетов")
@allure.story("Создание шаблона")
@allure.title("Yangi Шаблон отчета yaratish (.xlsx fayl bilan)")
def test_report_template_create(page: Page, code) -> None:
    with allure.step("Tizimga kirish"):
        authorization(page)
    run_report_template(page, code)


# ---------------------------------------------------------------------------------------------------


def run_report_template_edit(page: Page, code) -> None:
    """Yangi shablon yaratib, Изменить orqali Название'ni o'zgartiradi (fayl saqlanadi,
    qayta yuklanmaydi) va o'zgarish ro'yxatda aks etganini tekshiradi."""
    m = BasePage(page)
    old_name = f"shablon-edit-{code}"
    new_name = f"shablon-upd-{code}"

    run_report_template(page, code, name=old_name)

    with allure.step(f"'{old_name}' → Изменить"):
        m.click_grid_row(old_name)
        m.click_button("Изменить")
        # Create/edit heading bir xil ("Настройки шаблонов") — forma yuklanganini
        # Название qiymatidan bilamiz (company_client patterni)
        m.input(label="Название", expect_value=old_name)

    with allure.step(f"Tahrirlash: {old_name} → {new_name}"):
        m.input(label="Название", value=new_name)
        m.save()

    with allure.step(f"Ro'yxatda '{new_name}' bor, '{old_name}' YO'Q"):
        flow_navigate(page, tab="Модератор", name=MENU, expect_url="template_list")
        m.expect_heading(MENU)
        m.search(new_name)
        m.grid_row(new_name)
        m.search(old_name)
        m.expect_no_row(old_name)


@allure.epic("Модератор")
@allure.feature("Шаблоны отчетов")
@allure.story("Редактирование шаблона")
@allure.title("Шаблон отчетаni tahrirlash va o'zgarishni tekshirish")
def test_report_template_edit(page: Page, code) -> None:
    with allure.step("Tizimga kirish"):
        authorization(page)
    run_report_template_edit(page, code)


# ---------------------------------------------------------------------------------------------------


def run_report_template_status(page: Page, code) -> None:
    """Status ikki yo'nalishda — Изменить formasidagi Статус switch orqali:
    Активный → Неактивный → qayta Активный. DIQQAT: bu ro'yxat passive yozuvni
    YASHIRMAYDI (active+passive birga ko'rinadi) — status faqat grid badge orqali
    tekshiriladi."""
    m = BasePage(page)
    name = f"shablon-stat-{code}"

    run_report_template(page, code, name=name)

    with allure.step(f"1) '{name}' → Изменить → Неактивный"):
        m.click_grid_row(name)
        m.click_button("Изменить")
        m.input(label="Название", expect_value=name)
        m.checkbox(label="Статус", checked=False)
        m.save()

    with allure.step("1) Ro'yxatда Неактивный (passive) bo'lib ko'rinishi"):
        # DIQQAT: bu ro'yxat passive yozuvni YASHIRMAYDI (active+passive birga) —
        # show_all KERAK EMAS; faqat grid badge tekshiriladi.
        flow_navigate(page, tab="Модератор", name=MENU, expect_url="template_list")
        m.expect_heading(MENU)
        m.search(name)
        m.grid_row(name, "Неактивный")

    with allure.step(f"2) '{name}' → Изменить → Активный"):
        m.click_grid_row(name)
        m.click_button("Изменить")
        m.checkbox(label="Статус", checked=True)
        m.save()

    with allure.step("2) Ro'yxatda Активный bo'lib ko'rinishi"):
        flow_navigate(page, tab="Модератор", name=MENU, expect_url="template_list")
        m.expect_heading(MENU)
        m.search(name)
        m.grid_row(name, "Активный")


@allure.epic("Модератор")
@allure.feature("Шаблоны отчетов")
@allure.story("Статус шаблона")
@allure.title("Шаблон отчета statusini Неактивный/Активный qilish")
def test_report_template_status(page: Page, code) -> None:
    with allure.step("Tizimga kirish"):
        authorization(page)
    run_report_template_status(page, code)


# ---------------------------------------------------------------------------------------------------


def run_report_template_delete(page: Page, code) -> None:
    """Yangi shablon yaratib, Удалить orqali o'chiradi (cdk "Да/Нет") va ro'yxatdan
    yo'qolganini tekshiradi."""
    m = BasePage(page)
    name = f"shablon-del-{code}"

    run_report_template(page, code, name=name)

    with allure.step(f"'{name}' → Удалить → 'да'"):
        m.click_grid_row(name)
        m.click_button("Удалить")
        m.confirm("да")

    with allure.step(f"'{name}' ro'yxatdan yo'qolganini tekshirish"):
        m.search(name)
        m.expect_no_row(name)


@allure.epic("Модератор")
@allure.feature("Шаблоны отчетов")
@allure.story("Удаление шаблона")
@allure.title("Шаблон отчетаni o'chirish va ro'yxatdan yo'qolganini tekshirish")
def test_report_template_delete(page: Page, code) -> None:
    with allure.step("Tizimga kirish"):
        authorization(page)
    run_report_template_delete(page, code)


# ---------------------------------------------------------------------------------------------------


def run_report_template_negative(page: Page) -> None:
    """Negativ: bo'sh forma (Источник/Название/Файл to'ldirilmagan) + Сохранить →
    save BLOKLANADI, forma ochiq qoladi (majburiy maydonlar tekshiruvi)."""
    m = BasePage(page)

    with allure.step(f"Навигация → Создать: {MENU}"):
        _open_create(page, m)

    with allure.step("Bo'sh forma bilan Сохранить → bloklanishi kerak"):
        m.click_button("Сохранить")
        m.wait_for_loader()
        # Save bloklansa forma ochiq qoladi: URL hali setting-formasida (list'ga ketmaydi)
        expect(page).to_have_url(re.compile(r"setting"))
        m.expect_heading(FORM_HEADING)


@allure.epic("Модератор")
@allure.feature("Шаблоны отчетов")
@allure.story("Негатив — обязательные поля")
@allure.title("Bo'sh forma bilan saqlash bloklanishi")
def test_report_template_negative(page: Page) -> None:
    with allure.step("Tizimga kirish"):
        authorization(page)
    run_report_template_negative(page)
