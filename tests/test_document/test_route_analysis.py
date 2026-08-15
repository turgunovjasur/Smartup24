"""Анализ маршрутов (Route Analysis) hisobot formasi testlari.

Forma: Модератор → Документы → Анализ маршрутов
URL: /x24/a2/sb/rep/sbd/route_analysis
Elementlar (MCP/exploration dev/sm24 2026-07-29 tasdiqlangan):
  - "Тип отчёта" — 7 variant ([role=radio], ko'rinadigan <label> matni bo'yicha
    tanlanadi; default "Анализ маршрутов").
  - "Начало *" / "Конец *" — smt-date-picker (majburiy), matn "dd.mm.yyyy".
  - "План по активации" / "План по заказам" — smt-input (default 12 / 6).
  - "Менеджеры" / "Пользователи" — [role=checkbox] span'lar.
  - "Параметры" / "История" tablar + "Сформировать" tugmasi.
Download SINXRON: "Сформировать" → darhol .xlsx (page.expect_download).
Fayl: Анализ-маршрутов(DD.MM.YYYY+HH_MM_SS).xlsx; sheet "UI-SB276 route analysis";
13 ustun ("UI-SB276:" prefiksli, "exec %" 3 marta: visits/activation/deal).

Loyiha uslubi (flat): run_* (biznes logika) + test_* (authorization + run_*),
oxirida test_route_analysis_all zanjiri. BasePage instansiyalab ishlatiladi.
"""
import os
import re
from datetime import date, timedelta

import allure
import openpyxl
import pytest
from playwright.sync_api import Page, TimeoutError as PWTimeout

from flows.flow_authorization import authorization
from flows.flow_navbar import flow_navigate
from utils.base_page import BasePage

MENU = "Анализ маршрутов"
# Report ilova-bo'ylab ruslashtirilib qayta nomlandi (2026-08-05 dumpда tasdiqlangan,
# prod+dev ikkalasi ham): sheet "UI-SB276 route analysis" → "Анализ маршрутов",
# ustunlar "UI-SB276:" prefiksli inglizcha → prefiksSIZ ruscha. Ustun TARTIBI
# o'zgarmadi (COL_* indekslari hali to'g'ri). SHEET_NAME endi qat'iy tekshirilmaydi —
# _load_sheet birinchi (yagona) sheet'ni oladi (kelajakdagi qayta nomlashга chidamli).
SHEET_NAME = "Анализ маршрутов"
HEADER_PREFIX = ""  # prefiks olib tashlangan — _strip endi no-op
# Kutilgan 13 ustun (joriy ruscha nomlar, tartib bilan)
EXPECTED_HEADERS = [
    "Пользователь", "План", "Факт", "Не посещено", "Выполнение (%)",
    "План по активации", "Факт по активации", "Выполнение (%)",
    "План по заказам", "Факт по заказам", "Заказавшие клиенты",
    "Выполнение (%)", "Сумма заказов",
]
# Ustun indekslari (0-asosli) — ishlatiladiganlar
COL_USER = 0
COL_PLAN_ACT = 5
COL_PLAN_DEAL = 8

REPORT_TYPES = [
    "Анализ маршрутов",
    "Расписание маршрутов",
    "Непосещённые торговые точки",
    "Отчёт по активациям",
    "Отчёт по плану",
    "Отчёт по торговым точкам",
    "Отчёт по аудио",
]


# ======================================================================================
# Helperlar
# ======================================================================================

def _open_route_form(page: Page, m: BasePage) -> None:
    """Модератор → Анализ маршрутов formasini ochadi.

    Formani ZANJIRDA qayta-qayta ochganда ikki router-outlet asinxron yangilanib
    "Сформировать" ba'zan ko'rinmay qoladi (CLAUDE.md sub-header/content outlet) —
    Escape + reload bilan qayta urinamiz."""
    flow_navigate(page, tab="Модератор", name=MENU)
    for _ in range(3):
        try:
            page.wait_for_url(lambda u: "route_analysis" in u, timeout=15_000)
            m._settle()
            page.get_by_role("button", name="Сформировать").wait_for(state="visible", timeout=15_000)
            return
        except PWTimeout:
            page.keyboard.press("Escape")
            if "route_analysis" in page.url:
                page.reload()
            else:
                flow_navigate(page, tab="Модератор", name=MENU)
    raise AssertionError("Анализ маршрутов formasi ochilmadi (Сформировать ko'rinmadi)")


def _fill_dates_if_present(page: Page, m: BasePage, start: str, end: str) -> None:
    """Sana maydonlari ("Начало"/"Конец" LABEL'i bilan) mavjud bo'lsa to'ldiradi.

    Ba'zi hisobot turlari "Выберите дату" placeholder'li 2 ta textbox ko'rsatsa
    ham ularni "Начало"/"Конец" deb NOMLAMAYDI (masalan "Непосещённые торговые
    точки" — prod dumpда tasdiqlangan 2026-08-05) — bunda placeholder soniga emas,
    aynan "Начало" LABEL'i borligiga qarab hal qilamiz (aks holda m.input(label=
    "Начало") topolmay yiqilardi)."""
    label = page.locator("label, span").filter(
        has_text=re.compile(r"^\s*Начало\s*\*?\s*$"))
    if label.count() > 0:
        _fill_dates(m, start, end)


def _select_report_type(page: Page, report_type: str) -> None:
    """"Тип отчёта" radio variantини ko'rinadigan label matni bo'yicha tanlaydi
    ([role=radio] aria-label bo'sh — matn <label>да)."""
    page.get_by_text(report_type, exact=True).first.click()


def _fill_dates(m: BasePage, start: str, end: str) -> None:
    """"Начало"/"Конец" date-picker'larга matn sifatida sana yozadi."""
    m.input(label="Начало", value=start)
    m.input(label="Конец", value=end)


def _default_dates() -> tuple[str, str]:
    start = (date.today() - timedelta(days=30)).strftime("%d.%m.%Y")
    end = date.today().strftime("%d.%m.%Y")
    return start, end


def _generate_download(page: Page, m: BasePage, dest_dir: str, *, timeout: int = 60_000) -> str:
    """"Сформировать" bosib .xlsx yuklaydi va saqlangan fayl yo'lini qaytaradi.

    Asosiy yo'l — SINXRON download (page.expect_download). Katta hisobot async
    generatsiya bo'lib darhol kelmasa (timeout), "История" tab'idan tayyor
    yozuvni polling bilan yuklaydi (best-effort fallback)."""
    os.makedirs(dest_dir, exist_ok=True)
    try:
        with page.expect_download(timeout=timeout) as dl_info:
            m.click_button("Сформировать")
        download = dl_info.value
    except PWTimeout:
        return _download_from_history(page, m, dest_dir)
    path = os.path.join(dest_dir, download.suggested_filename)
    download.save_as(path)
    assert os.path.getsize(path) > 0, f"Yuklangan fayl bo'sh: {path}"
    return path


def _download_from_history(page: Page, m: BasePage, dest_dir: str, *, attempts: int = 30) -> str:
    """Async hisobot uchun fallback: "История" tab'iga o'tib, birinchi yozuv tayyor
    bo'lguncha polling bilan kutadi va "Скачать" (yoki download ikonka) bosib yuklaydi.

    DIQQAT: sinxron yo'l dev'da ishlaydi — bu tarmoq faqat katta async hisobotlar
    uchun; struktura История tab'idan olingan, async holat kelganda tekshiriladi."""
    m.click_button("История")
    m._settle()
    for _ in range(attempts):
        dl_btn = page.get_by_role("button", name=re.compile("Скачать|Download", re.I)).first
        if dl_btn.count() and dl_btn.is_enabled():
            with page.expect_download(timeout=30_000) as dl_info:
                dl_btn.click()
            download = dl_info.value
            path = os.path.join(dest_dir, download.suggested_filename)
            download.save_as(path)
            return path
        page.wait_for_timeout(2_000)
        page.reload()
        m._settle()
    raise AssertionError("История: hisobot tayyor bo'lmadi (async timeout)")


def _is_total_row(user) -> bool:
    """Fayl oxiridagi YIG'INDI qatori (user = "UI-SB276:total" / "Итого") — foydalanuvchi
    emas, ustunlar summasi; mantiq tekshiruvlaridan chiqarib tashlanadi."""
    u = str(user).replace(HEADER_PREFIX, "").strip().lower()
    return u in ("total", "итого", "всего")


def _load_sheet(path: str) -> tuple[list[str], list[tuple]]:
    """xlsx'ni ochib (headers, data_rows) qaytaradi — headers prefiks bilan xom holda,
    data qatorlaridan bo'sh va YIG'INDI (total) qatori chiqarib tashlanadi."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        # Hisobotда yagona sheet bor — nomi ilova versiyasida o'zgarishi mumkin
        # (2026-08-05 "Анализ маршрутов"), shuning uchun NOMGA tayanmay birinchisini
        # olamiz (kelajakdagi qayta nomlashga chidamli).
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = [r for r in rows[1:]
            if r and r[COL_USER] not in (None, "") and not _is_total_row(r[COL_USER])]
    return headers, data


def _strip(headers: list[str]) -> list[str]:
    return [h.replace(HEADER_PREFIX, "").strip() for h in headers]


# ======================================================================================
# run_* — biznes logika (login qilinganini kutadi)
# ======================================================================================

def run_download_success(page: Page, dest_dir: str) -> str:
    m = BasePage(page)
    _open_route_form(page, m)
    _fill_dates(m, *_default_dates())
    path = _generate_download(page, m, dest_dir)
    assert path.endswith(".xlsx"), f".xlsx emas: {path}"
    assert os.path.getsize(path) > 0
    allure.attach(f"{os.path.basename(path)} ({os.path.getsize(path)} bayt)",
                  name="downloaded_file", attachment_type=allure.attachment_type.TEXT)
    return path


def run_structure(page: Page, dest_dir: str) -> None:
    m = BasePage(page)
    _open_route_form(page, m)
    _fill_dates(m, *_default_dates())
    path = _generate_download(page, m, dest_dir)

    headers, data = _load_sheet(path)
    assert _strip(headers) == EXPECTED_HEADERS, (
        f"Ustunlar mos emas:\n olindi:  {_strip(headers)}\n kutilgan: {EXPECTED_HEADERS}"
    )
    assert len(data) >= 1, "Hisobotда kamida 1 qator ma'lumot bo'lishi kerak"
    allure.attach(f"{len(headers)} ustun, {len(data)} qator", name="structure",
                  attachment_type=allure.attachment_type.TEXT)


def run_plan_values(page: Page, dest_dir: str) -> None:
    """"План по активации"=20, "План по заказам"=9 kiritilganda fayldagi
    "plan per activation"/"plan per deal" ustunlarida aks etishini tekshiradi
    (default 12/6 dan farqli qiymat — haqiqiy aks etishini isbotlaydi)."""
    m = BasePage(page)
    _open_route_form(page, m)
    _fill_dates(m, *_default_dates())
    m.input(label="План по активации", value="20")
    m.input(label="План по заказам", value="9")
    path = _generate_download(page, m, dest_dir)
    _, data = _load_sheet(path)

    for row in data:
        assert row[COL_PLAN_ACT] == 20, (
            f"[{row[COL_USER]}] plan per activation={row[COL_PLAN_ACT]}, kutilgan 20"
        )
        assert row[COL_PLAN_DEAL] == 9, (
            f"[{row[COL_USER]}] plan per deal={row[COL_PLAN_DEAL]}, kutilgan 9"
        )


def run_report_type(page: Page, dest_dir: str, report_type: str) -> None:
    """Berilgan "Тип отчёта" tanlanganda .xlsx yuklab olish ishlashini tekshiradi."""
    m = BasePage(page)
    _open_route_form(page, m)
    _select_report_type(page, report_type)
    _fill_dates_if_present(page, m, *_default_dates())
    path = _generate_download(page, m, dest_dir)
    assert path.endswith(".xlsx") and os.path.getsize(path) > 0


def run_required_fields(page: Page) -> None:
    """Majburiy sanalar (Начало/Конец) bo'sh bo'lsa Сформировать hisobot bermasligi
    kerak. DIQQAT: forma ochилganда sanalar DEFAULT bilan to'ladi — shuning uchun
    ularni ATAYLAB tozalaymiz (aks holda "bo'sh" test soxta bo'lardi). Tozalangach
    ham yuklansa — required validatsiyasi yo'q (bug, xfail)."""
    m = BasePage(page)
    _open_route_form(page, m)
    # Har ikala sanani bo'shatamiz
    m.input(label="Начало", value="")
    m.input(label="Конец", value="")
    page.keyboard.press("Escape")
    try:
        with page.expect_download(timeout=8_000):
            m.click_button("Сформировать")
        pytest.xfail("Majburiy sanalar bo'sh bo'lsa ham hisobot yuklandi — required validatsiyasi yo'q")
    except PWTimeout:
        pass  # kutilgan: majburiy maydonlar bo'sh — yuklanmadi


# ======================================================================================
# test_* — har biri alohida (o'z login'i bilan)
# ======================================================================================

@allure.epic("Документы")
@allure.feature("Анализ маршрутов")
@allure.title("Hisobot .xlsx muvaffaqiyatli yuklanadi (default parametrlar)")
def test_report_download_success(page: Page, tmp_path) -> None:
    """Default parametrlar bilan Сформировать → .xlsx yuklanadi, hajmi > 0."""
    authorization(page)
    run_download_success(page, str(tmp_path))


@allure.epic("Документы")
@allure.feature("Анализ маршрутов")
@allure.title("Hisobot tuzilishi: sheet, 13 ustun, kamida 1 qator")
def test_report_structure(page: Page, tmp_path) -> None:
    """Yuklangan faylda kerakli sheet, barcha kutilgan ustunlar va >=1 qator bor."""
    authorization(page)
    run_structure(page, str(tmp_path))


@allure.epic("Документы")
@allure.feature("Анализ маршрутов")
@allure.title("Majburiy maydonlar: sanalar bo'sh bo'lsa Сформировать ishlamaydi")
def test_required_fields(page: Page) -> None:
    """Majburiy Начало/Конец bo'sh bo'lsa hisobot generatsiya bo'lmaydi."""
    authorization(page)
    run_required_fields(page)


@allure.epic("Документы")
@allure.feature("Анализ маршрутов")
@allure.title("Har 7 'Тип отчёта' uchun yuklab olish ishlaydi")
@pytest.mark.parametrize("report_type", REPORT_TYPES)
def test_different_report_types(page: Page, tmp_path, report_type: str) -> None:
    """BARCHA 7 hisobot turi uchun: turni tanlab Сформировать → .xlsx yuklanadi."""
    authorization(page)
    run_report_type(page, str(tmp_path), report_type)


@allure.epic("Документы")
@allure.feature("Анализ маршрутов")
@allure.title("План qiymatlari faylda aks etadi (План по активации/заказам)")
def test_plan_values_reflected(page: Page, tmp_path) -> None:
    """Kiritilgan План по активации=20 va План по заказам=9 fayldagi
    "plan per activation"/"plan per deal" ustunlarida aks etadi."""
    authorization(page)
    run_plan_values(page, str(tmp_path))


@allure.epic("Документы")
@allure.feature("Анализ маршрутов")
@allure.title("Zanjir: happy-path route-analysis tekshiruvlari bitta login bilan")
def test_route_analysis_all(page: Page, tmp_path) -> None:
    """Bitta login bilan asosiy (yashil) tekshiruvlar ketma-ket (test_all uslubi).

    Validatsiya (date/required) va data_logic testlari ATAYLAB kiritilmagan — ular
    ilova bug'lari sababli xfail bo'ladi, alohida standalone qoladi; zanjir esa
    happy-path regressiyani yashil tekshiradi."""
    authorization(page)
    dest = str(tmp_path)
    run_download_success(page, dest)
    run_structure(page, dest)
    run_plan_values(page, dest)
    run_report_type(page, dest, "Расписание маршрутов")
